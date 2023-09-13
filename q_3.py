from q_1_mirror_class import Mirrors
from q_1_mirror_class import get_ord
import test
from random import random
import math
import matplotlib.pyplot as plt
import openpyxl

path="/Users/lvangge/Desktop/附件.xlsx"
r_1=100#内圈半径
r_2=350#外圈半径
r_3=5#相邻镜间的过道最小宽度
h_mir=4#镜子高度
a_q2=7.08#q2中模拟的镜子长度
b_q2=7.14#q2中模拟的镜子宽度
lr_q2=0.37#q2中模拟的径向扰动
lv_q2=0.29#q2中模拟的切向扰动
max_n=25#同心圆个数的一个上界
x,y=get_ord(path)
mir_before=Mirrors(x,y)
dv_before=mir_before.d_r()
dr_before=mir_before.d_v()['d']
dis_max=8#宽度最大值

def func(a,b,lr,lv):
    '''目标优化函数'''
    #生成一维坐标
    x_one=sum(ord_of_mir(b,lr,lv)['x'],[])
    y_one=sum(ord_of_mir(b,lr,lv)['y'],[])
    mlist=[test.Mirror(x_one[i],y_one[i]) for i in range(len(x_one))]
    f=test.E_mean(mlist)
  
    return f

def E_all(a,b,lr,lv):
    E_mean=func(a,b,lr,lv)
    E=0
    for i in range(len(num_of_mir(b,lr,lv))):
        E+=num_of_mir(b,lr,lv)[i]*a[i]*b[i]*E_mean
    return E

def get_d(b,lr,lv):
    '''得到定日镜排列方式  从径向和切向两方向'''
    dr,dv=[],[]
    i=0
    while sum(dr)<r_2-r_1:
        dr.append(lr[i]+r_3+b[i])
        i+=1
    dr.pop()
    dv=[lv[i]+r_3+b[i] for i in range(len(dr)+1)]
    
    return {'dr':dr,'dv':dv}
    #切向距离
    # dv=[
    #     mir_before.d_r()[k]+b-mir_before.b-random()+1/2 for k in range(len(mir_before.d_r()))
    # ]
    # #径向距离
    # dr=[
    #     mir_before.d_v()['d'][k]+b-mir_before.b-random()+1/2 for k in range(len(mir_before.d_v()['d']))
    # ]
    #考虑根据情况增加或者减少层数
    # dv=[
    #     dv_before[k]+b-mir_before.b for k in range(len(dv_before))
    # ]
    # #径向距离
    # dr=[
    #     dr_before[k]+b-mir_before.b for k in range(len(dr_before))
    # ]
    # cnt=0
    
    # #调增dr
    # if sum(dr)<r_2-r_1:
    #     while sum(dr)<r_2-r_1:
    #         dr.append(dr[-1])
    #         cnt+=1
    #     dr.pop()
    # elif sum(dr)>r_2-r_1:
    #     while sum(dr)>r_2-r_1:
    #         dr.pop()
    #         cnt-=1
    #     dr.append(dr[-1])
    # #调整dv
    # if cnt>0:
    #     for i in range(cnt-1):
    #         dv.append(dv[-1])
    # elif cnt<0:
    #     for i in range(1-cnt):
    #         dv.pop()
    

def s_of_mir(a,b):
    '''计算镜子面积'''
    return a*b

def distance(b,lr,lv):
    '''第i层的镜距中心集热柱的距离'''
    return [sum(get_d(b,lr,lv)['dr'][:i])+r_1 for i in range(len(get_d(b,lr,lv)['dr'])+1)]

def num_of_mir(b,lr,lv):
    '''计算第i层的镜子数量'''
    num=[]
    dv=get_d(b,lr,lv)['dv']
    for i in range(len(dv)):
        if i==0:
            r=r_1
        else:
            r=2*math.pi*distance(b,lr,lv)[i]
        num.append(int(r//dv[i]))
    
    return num

def ord_of_mir(b,lr,lv):
    '''获取集日镜的xy坐标 以二维矩阵输出 第i层第j个元素'''
    angle=[2 * math.pi / num_of_mir(b,lr,lv)[i] for i in range(len(num_of_mir(b,lr,lv)))]#第i层相邻两元素的夹角
    x=[
        [distance(b,lr,lv)[i] * math.cos(j*angle[i]) for j in range(num_of_mir(b,lr,lv)[i])]for i in range(len(num_of_mir(b,lr,lv)))
    ]
    y=[
        [distance(b,lr,lv)[i] * math.sin(j*angle[i]) for j in range(num_of_mir(b,lr,lv)[i])]for i in range(len(num_of_mir(b,lr,lv)))
    ]
    point={'x':x,'y':y}
    return point


class SA:
    '''退火算法'''
    def __init__(self,func,iter=100,T_max=1,T_min=0.01,alpha=0.9,h=h_mir) -> None:
        self.func=func
        self.iter=iter
        self.T_max=T_max
        self.T_min=T_min
        self.alpha=alpha
        self.T=T_max#当前温度
        #随机生成iter个目标点
        # #集日镜的长调整小于0.8
        self.a=[[(random()-random())*0.4+a_q2 for i in range(max_n) ] for i in range(iter)]
        #集日镜的宽等与长
        self.b=self.a
        #集日镜的高度
        self.h=h
        #径向方向调整<0.2
        self.lr=[[ (random()-random())*0.1+lr_q2 for i in range (max_n) ]for i in range(iter)]
        #切向方向调整<0.2
        self.lv=[[ (random()-random())*0.1+lv_q2 for i in range (max_n) ]for i in range(iter)]
        self.most_best=[]
        self.history={'f':[],'T':[]}
    
        
    def generatr_new(self,a,b,lr,lv):
        '''用于更新数据 严格落入[-5,5]定义域中'''
        while True:
            a_new=[a[i]+self.T*(random() - random()) for i in range(max_n)]
            b_new=a_new
            lr_new=[lr[i]+self.T*(random() - random())*0.1 for i in range(max_n)]
            lv_new=[lv[i]+self.T*(random() - random())*0.1 for i in range(max_n)]
            
            E_new=E_all(a_new,b_new,lr_new,lv_new)
            if (E_new>60) & all(x>0 for x in lv) & all(y>0 for y in lr):
                break
        return a_new,b_new,lr_new,lv_new
    
    def Metroplics(self,f,f_new):
        '''Metroplics准则'''
        if f<=f_new:
            return 1 
        else:
            p=math.exp((f_new - f) / self.T)
            if random() < p:
                return 1
            else:
                return 0
    
    def best(self):
        '''用于得出循环中各点中的最小值'''
        f_list=[]
        for i in range(self.iter):
            f_list.append(self.func(self.a[i],self.b[i],self.lr[i],self.lv[i]))
        f_best=max(f_list)
        ind=f_list.index(f_best)
        return f_best,ind
    

        
    def run(self):     
        '''运行SA'''
        count=0
        while self.T>self.T_min:#执行外循环
            
            for i in range(self.iter):#执行内循环
                f=self.func(self.a[i],self.b[i],self.lr[i],self.lv[i])
                a_new,b_new,lr_new,lv_new=self.generatr_new(self.a[i],self.b[i],self.lr[i],self.lv[i])
                f_new=self.func(a_new,b_new,lr_new,lv_new)
                if self.Metroplics(f,f_new):
                    self.a[i]=a_new
                    self.b[i]=b_new
                    self.lr[i]=lr_new
                    self.lv[i]=lv_new
            
            f_inside,ind_inside=self.best()
            self.history['f'].append(f_inside)
            self.history['T'].append(self.T)
            self.T=self.T * self.alpha
            count+=1
        
        f_final,ind_final=self.best()
        print(
            f"F={f_final},a={self.a[ind_final]},b={self.b[ind_final]},lr={self.lr[ind_final]},lv={self.lv[ind_final]}"
        )
        print(f"num{num_of_mir(self.b[ind_final],self.lr[ind_final],self.lv[ind_final])}")
        x_one=sum(ord_of_mir(self.b[ind_final],self.lr[ind_final],self.lv[ind_final])['x'],[])
        y_one=sum(ord_of_mir(self.b[ind_final],self.lr[ind_final],self.lv[ind_final])['y'],[])
        
        #总功率
        E=E_all(self.a[ind_final],self.b[ind_final],self.lr[ind_final],self.lv[ind_final])
        print(E)
        
        n=len(num_of_mir(self.b[ind_final],self.lr[ind_final],self.lv[ind_final]))
        del self.a[ind_final][n:]
        del self.b[ind_final][n:]
        del self.lr[ind_final][n:]
        del self.lv[ind_final][n:]
        
        #保存数据
        workbook1=openpyxl.Workbook()
        sheet1=workbook1.active
        data=[[x_one[i],y_one[i]] for i in range(len(x_one))]
        for row in data:
            sheet1.append(row)
        new_column_title = '长宽a'
        sheet1.cell(row=1, column=sheet1.max_column + 1, value=new_column_title)
        for i, value in enumerate(self.a[ind_final], start=2):
            sheet1.cell(row=i, column=sheet1.max_column, value=value)
            
        new_column_title = 'lr'
        sheet1.cell(row=1, column=sheet1.max_column + 1, value=new_column_title)
        for i, value in enumerate(self.lr[ind_final], start=2):
            sheet1.cell(row=i, column=sheet1.max_column, value=value)
          
        new_column_title = 'lv'
        sheet1.cell(row=1, column=sheet1.max_column + 1, value=new_column_title)
        for i, value in enumerate(self.lv[ind_final], start=2):
            sheet1.cell(row=i, column=sheet1.max_column, value=value)
       
        workbook1.save(filename='data3.xlsx')
        
        #画图
        plt.scatter(x_one,y_one,s=10,color='silver')
        plt.title("mirrors")
        plt.xlabel('x')
        plt.ylabel('y')
        plt.show()
        

sa=SA(func)
sa.run()


    
